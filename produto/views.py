import json
import os
import re
from datetime import timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO

import pandas as pd
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import F, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.deprecation import MiddlewareMixin
from django.views.decorators.http import require_POST
from django.views.generic.edit import CreateView, UpdateView
from openpyxl import load_workbook
from openpyxl.styles import (Alignment, Border, Font, NamedStyle, PatternFill,
                             Side)

from pdv.models import Compra, ItemCompra
from produto.forms import ProdutoForm, UploadFileForm
from produto.models import Categoria, Produto


class NavigationHistoryMiddleware(MiddlewareMixin):
    def process_request(self, request):
        if "navigation_history" not in request.session:
            request.session["navigation_history"] = []

        # Não adiciona POST nem URLs de edição ao histórico
        if (
            request.method == "GET"
            and not request.path.startswith("/produtos/")
            or not any(k in request.path for k in ["edit", "add"])
        ):
            history = request.session["navigation_history"]

            # Evita duplicatas consecutivas
            if not history or history[-1] != request.get_full_path():
                # Mantém apenas os últimos 5 itens
                if len(history) >= 5:
                    history.pop(0)
                history.append(request.get_full_path())
                request.session["navigation_history"] = history


@login_required
def index(request):
    template_name = "produto/index.html"
    query = request.GET.get("q")
    objects = Produto.objects.all()

    if query:
        queries = query.split()
        q_objects = Q()
        for q in queries:
            q_objects |= (
                Q(produto__icontains=q)
                | Q(codigoBarra__icontains=q)
                | Q(preco_venda__icontains=q)
                | Q(categoria__categoria__icontains=q)
            )
        objects = objects.filter(q_objects)

    objects = objects.order_by("produto")
    paginator = Paginator(objects, 50)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    if "list_origin" not in request.session:
        request.session["list_origin"] = request.get_full_path()

    # Adicionado para popular o <select> no modal de edição em massa
    todas_as_categorias = Categoria.objects.all()

    context = {
        "page_obj": page_obj,
        "todas_as_categorias": todas_as_categorias,  # Passando para o template
    }
    return render(request, template_name, context)


def product_detail(request, pk):
    if "detail_origin" not in request.session:
        request.session["detail_origin"] = request.META.get(
            "HTTP_REFERER", reverse("produto:index")
        )

    template_name = "produto/product_detail.html"
    obj = get_object_or_404(Produto, pk=pk)
    context = {"object": obj}
    return render(request, template_name, context)


@login_required
def purchase_details(request, pk):
    compra = get_object_or_404(Compra, pk=pk)
    context = {"compra": compra, "itens": compra.itens.all()}
    return render(request, "produto/purchase_details.html", context)


@login_required
def product_add(request):
    template_name = "produto/product_form.html"
    return render(request, template_name)


class ProductCreate(LoginRequiredMixin, CreateView):
    model = Produto
    template_name = "produto/product_form.html"
    form_class = ProdutoForm
    login_url = "/login"

    def form_valid(self, form):
        response = super().form_valid(form)
        if not self.object.codigoBarra:
            # Atualiza o código de barras com o ID do produto gerado após o primeiro salvamento
            self.object.codigoBarra = str(self.object.pk)
            self.object.save(update_fields=["codigoBarra"])
        return response


class ProdutoUpdate(LoginRequiredMixin, UpdateView):
    model = Produto
    template_name = "produto/product_form.html"
    form_class = ProdutoForm
    login_url = "/login"

    def get(self, request, *args, **kwargs):
        # Armazena origem da edição (página de detalhes)
        request.session["edit_origin"] = request.META.get(
            "HTTP_REFERER", reverse("produto:index")
        )
        return super().get(request, *args, **kwargs)

    def get_success_url(self):
        # Volta para a origem da edição (página de detalhes)
        return self.request.session.get(
            "edit_origin",
            reverse("produto:product_detail", kwargs={"pk": self.object.pk}),
        )

    def get_success_url(self):
        # Redireciona para a página de detalhes do produto após a atualização
        return reverse("produto:product_detail", kwargs={"pk": self.object.pk})


@login_required
def get_products_data(request):
    product_ids_str = request.GET.get("ids")
    if not product_ids_str:
        return JsonResponse({"error": "Nenhum ID fornecido"}, status=400)

    product_ids = [int(id) for id in product_ids_str.split(",")]

    # Usamos values() para buscar apenas os campos que precisamos de forma eficiente
    products = Produto.objects.filter(id__in=product_ids).values(
        "id", "produto", "preco_custo", "preco_venda", "estoque"
    )

    # Convertemos o QuerySet para uma lista de dicionários
    data = list(products)

    return JsonResponse(data, safe=False)


@login_required
@require_POST
def bulk_edit_products(request):
    try:
        # Carregamos os dados do corpo da requisição JSON
        data = json.loads(request.body)

        if not isinstance(data, list) or not data:
            return JsonResponse(
                {"status": "error", "message": "Dados inválidos ou vazios."}, status=400
            )

        with transaction.atomic():
            # Itera sobre cada produto enviado no JSON
            for product_data in data:
                product_id = product_data.get("id")

                # Encontra o produto no banco
                produto_obj = Produto.objects.get(pk=product_id)

                # Atualiza os campos se eles foram fornecidos e não estão vazios
                # Usamos .get() com um valor padrão para evitar erros se a chave não existir
                if "preco_custo" in product_data and product_data["preco_custo"] != "":
                    produto_obj.preco_custo = Decimal(product_data["preco_custo"])

                if "preco_venda" in product_data and product_data["preco_venda"] != "":
                    produto_obj.preco_venda = Decimal(product_data["preco_venda"])

                if "estoque" in product_data and product_data["estoque"] != "":
                    produto_obj.estoque = int(product_data["estoque"])

                # Salvamos o objeto individualmente para garantir que o método .save()
                # customizado do seu modelo (que recalcula margem e nível de estoque) seja executado.
                produto_obj.save()

        messages.success(
            request, f"{len(data)} produtos foram atualizados com sucesso!"
        )
        return JsonResponse({"status": "success", "message": "Produtos atualizados."})

    except json.JSONDecodeError:
        return JsonResponse(
            {"status": "error", "message": "JSON mal formatado."}, status=400
        )
    except Produto.DoesNotExist:
        return JsonResponse(
            {
                "status": "error",
                "message": f"Produto com ID {product_id} não encontrado.",
            },
            status=404,
        )
    except (KeyError, TypeError, ValueError) as e:
        return JsonResponse(
            {"status": "error", "message": f"Erro nos dados fornecidos: {str(e)}"},
            status=400,
        )
    except Exception as e:
        return JsonResponse(
            {"status": "error", "message": f"Ocorreu um erro inesperado: {str(e)}"},
            status=500,
        )


def import_xlsx(file_path):
    """
    Importa planilhas xlsx.
    """
    workbook = load_workbook(file_path)
    sheet = workbook.active

    categorias = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        categoria = row[8]
        if categoria:
            categorias.add(categoria)

    for categoria in categorias:
        Categoria.objects.get_or_create(categoria=categoria)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        nivel_estoque = row[0]
        produto_nome = row[1]

        try:
            preco_custo = (
                Decimal(str(row[2]).replace(",", "."))
                if row[2] not in (None, "")
                else Decimal("0.00")
            )
            preco_venda = (
                Decimal(str(row[3]).replace(",", "."))
                if row[3] not in (None, "")
                else Decimal("0.00")
            )
            margem_venda = (
                Decimal(str(row[4]).replace(",", "."))
                if row[4] not in (None, "")
                else Decimal("0.00")
            )
        except (ValueError, InvalidOperation):
            preco_custo = Decimal("0.00")
            preco_venda = Decimal("0.00")
            margem_venda = Decimal("0.00")

        try:
            estoque = (
                int(row[5]) if row[5] not in (None, "") and str(row[5]).isdigit() else 0
            )
            estoque_minimo = (
                int(row[6]) if row[6] not in (None, "") and str(row[6]).isdigit() else 0
            )
        except ValueError:
            estoque = 0
            estoque_minimo = 0

        codigoBarra = row[7] if row[7] not in (None, "") else ""
        categoria_nome = row[8] if row[8] not in (None, "") else ""

        categoria = Categoria.objects.filter(categoria=categoria_nome).first()

        produto_data = {
            "nivel_estoque": nivel_estoque,
            "preco_custo": preco_custo,
            "preco_venda": preco_venda,
            "margem_vendas": margem_venda,
            "estoque": estoque,
            "estoque_minimo": estoque_minimo,
            "codigoBarra": codigoBarra,
            "categoria": categoria,
        }

        Produto.objects.update_or_create(produto=produto_nome, defaults=produto_data)


@login_required
@require_POST
def import_data(request):
    form = UploadFileForm(request.POST, request.FILES)
    if form.is_valid():
        file = request.FILES["file"]
        file_path = f"/tmp/{file.name}"
        with open(file_path, "wb+") as destination:
            for chunk in file.chunks():
                destination.write(chunk)
        try:
            import_xlsx(file_path)
            messages.success(request, "Dados importados com sucesso.")
        except Exception as e:
            messages.error(request, f"Erro ao importar dados: {str(e)}")
        os.remove(file_path)
        return redirect("produto:index")
    else:
        messages.error(request, "Erro ao fazer upload do arquivo.")
    return redirect("produto:upload")


@login_required
def upload_file(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES["file"]
            if file.name.endswith(".xlsx"):
                return redirect("produto:import_data")
            else:
                messages.error(
                    request, "O arquivo enviado não é um arquivo Excel válido."
                )
    else:
        form = UploadFileForm()
    return render(request, "produto/upload.html", {"form": form})


@login_required
def gerar_insights(request):
    try:
        # Importar datetime no topo do arquivo, se ainda não estiver importado
        from datetime import datetime

        # Obter a data atual
        data_atual = datetime.now().strftime("%d-%m-%Y")

        # Obter todos os produtos
        df_produtos = pd.DataFrame(list(Produto.objects.all().values()))

        # Filtrar produtos onde preço de venda e estoque mínimo não estão zerados ou nulos
        df_produtos = df_produtos[
            (df_produtos["preco_venda"] > 0) & (df_produtos["estoque_minimo"] > 0)
        ]

        # Processar dados
        produtos_abaixo_estoque_minimo = df_produtos[
            df_produtos["estoque"] < df_produtos["estoque_minimo"]
        ]
        df_produtos["valor_venda_pot"] = (
            df_produtos["estoque"] * df_produtos["preco_venda"]
        )
        produtos_preco_zero = df_produtos[df_produtos["preco_venda"] == 0]
        produtos_estoque_excedente = df_produtos[
            df_produtos["estoque"] > df_produtos["estoque_minimo"] * 2
        ]
        df_produtos["rotatividade"] = (
            df_produtos["estoque"] / df_produtos["estoque_minimo"]
        )
        df_produtos["custo_estoque"] = (
            df_produtos["estoque"] * df_produtos["preco_custo"]
        )
        produtos_alto_custo_estoque = df_produtos[df_produtos["custo_estoque"] > 1000]

        # Obter todas as categorias e seus produtos
        categorias = Categoria.objects.all().prefetch_related("produto_set")

        # Calcular o valor total das vendas
        valor_total_vendas = (df_produtos["estoque"] * df_produtos["preco_venda"]).sum()

        # Obter todos os detalhes de pagamento
        df_compras = pd.DataFrame(
            list(Compra.objects.all().values("id", "data", "metodo_pagamento", "total"))
        )
        df_itens_compra = pd.DataFrame(
            list(
                ItemCompra.objects.all().values(
                    "compra_id", "produto_id", "quantidade", "preco_unitario"
                )
            )
        )

        # Obter o nome dos produtos
        df_itens_compra = df_itens_compra.merge(
            df_produtos[["id", "produto"]],
            left_on="produto_id",
            right_on="id",
            how="left",
        )

        # Converta datetimes para timezone naive e formate para o padrão brasileiro
        if not df_compras.empty:
            df_compras["data"] = pd.to_datetime(df_compras["data"]).dt.tz_localize(None)
            df_compras["data"] = df_compras["data"].dt.strftime("%d/%m/%Y %H:%M")

        # Adicionar detalhes de itens de compra
        df_compras_details = df_compras.merge(
            df_itens_compra, left_on="id", right_on="compra_id", how="left"
        )

        # Adicionar coluna de semana
        df_compras_details["semana"] = (
            df_compras_details["data"]
            .apply(lambda x: pd.to_datetime(x, format="%d/%m/%Y %H:%M"))
            .dt.to_period("W")
            .apply(lambda r: r.start_time.strftime("%d/%m/%Y %H:%M"))
        )

        # Agrupar por semana e somar os preços unitários
        df_totais_semanais = (
            df_compras_details.groupby("semana")["preco_unitario"].sum().reset_index()
        )
        df_totais_semanais.columns = ["Semana", "Total Preço Unitário"]

        # Substituir ponto por vírgula nas colunas numéricas especificadas
        for col in ["preco_custo", "preco_venda", "margem_vendas"]:
            df_produtos[col] = df_produtos[col].apply(
                lambda x: str(x).replace(".", ",")
            )

        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            # Exportar DataFrame `df_produtos` para Excel
            df_produtos.to_excel(writer, sheet_name="Produtos", index=False)

            def style_worksheet(worksheet):
                # Estilo para cabeçalhos
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(
                    start_color="4F81BD", end_color="4F81BD", fill_type="solid"
                )

                column_widths = [
                    max(len(str(cell.value)) for cell in col)
                    for col in worksheet.columns
                ]
                for i, width in enumerate(column_widths):
                    worksheet.column_dimensions[chr(65 + i)].width = width + 6

                for cell in worksheet[1]:  # Cabeçalhos
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )

                # Estilo para linhas
                border_style = Border(
                    left=Side(border_style="thin"),
                    right=Side(border_style="thin"),
                    top=Side(border_style="thin"),
                    bottom=Side(border_style="thin"),
                )
                fill_odd = PatternFill(
                    start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                )
                fill_even = PatternFill(
                    start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
                )

                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.border = border_style
                        cell.fill = fill_even if cell.row % 2 == 0 else fill_odd
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True
                        )

            # Aplicar estilos na aba "Produtos"
            style_worksheet(writer.sheets["Produtos"])

            # Nova aba: Produtos por Categoria (com validação)
            produtos_por_categoria = []
            for categoria in categorias:
                produtos = categoria.produto_set.filter(preco_venda__gt=0)
                for produto in produtos:
                    produtos_por_categoria.append(
                        {
                            "Categoria": categoria.categoria,
                            "Produto": produto.produto,
                            "Preço de Custo": produto.preco_custo,
                            "Preço de Venda": produto.preco_venda,
                            "Margem de Venda": produto.margem_vendas,
                            "Estoque": produto.estoque,
                            "Estoque Mínimo": produto.estoque_minimo,
                            "Código de Barras": produto.codigoBarra,
                            "Nível de Estoque": produto.nivel_estoque,
                        }
                    )

            df_produtos_por_categoria = pd.DataFrame(produtos_por_categoria)
            if not df_produtos_por_categoria.empty:
                # Substituição de '.' por ',' nas colunas numéricas
                for col in ["Preço de Custo", "Preço de Venda", "Margem de Venda"]:
                    df_produtos_por_categoria[col] = df_produtos_por_categoria[
                        col
                    ].apply(lambda x: str(x).replace(".", ","))

                df_produtos_por_categoria.to_excel(
                    writer, sheet_name="Produtos por Categoria", index=False
                )
            else:
                pd.DataFrame(
                    {
                        "Mensagem": [
                            "Nenhum produto atende aos critérios de preço de venda > 0 e estoque > 0"
                        ]
                    }
                ).to_excel(writer, sheet_name="Produtos por Categoria", index=False)

            # Adicionar abas com os dados
            produtos_abaixo_estoque_minimo.to_excel(
                writer, sheet_name="Abaixo Estoque", index=False
            )
            df_produtos[["produto", "valor_venda_pot"]].to_excel(
                writer, sheet_name="Venda Potencial", index=False
            )
            produtos_preco_zero.to_excel(writer, sheet_name="Preço Zerado", index=False)
            produtos_estoque_excedente.to_excel(
                writer, sheet_name="Estoque Excedente", index=False
            )
            df_produtos[["produto", "rotatividade"]].to_excel(
                writer, sheet_name="Rotatividade", index=False
            )
            produtos_alto_custo_estoque.to_excel(
                writer, sheet_name="Custo Alto Estoque", index=False
            )

            # Adicionar aba com o valor total das vendas
            total_vendas_df = pd.DataFrame({"Total Vendas": [valor_total_vendas]})
            total_vendas_df.to_excel(writer, sheet_name="Total Vendas", index=False)

            # Adicionar aba com os detalhes dos pagamentos
            df_compras_details[
                [
                    "data",
                    "metodo_pagamento",
                    "total",
                    "produto",
                    "quantidade",
                    "preco_unitario",
                ]
            ].to_excel(writer, sheet_name="Detalhes Pagamentos", index=False)

            # Adicionar aba com os totais semanais
            df_totais_semanais.to_excel(
                writer, sheet_name="Totais Semanais", index=False
            )

            # Aplicar formatação
            for sheet_name in writer.sheets:
                sheet = writer.sheets[sheet_name]
                style_worksheet(sheet)

                # Adicionar rodapé com totais (se necessário)
                if sheet_name == "Detalhes Pagamentos":
                    last_row = sheet.max_row + 1
                    sheet[f"A{last_row}"] = "Total por Semana"
                    for col in range(2, sheet.max_column + 1):
                        sheet.cell(row=last_row, column=col).value = (
                            "=SUM({}:{})".format(
                                sheet.cell(row=2, column=col).coordinate,
                                sheet.cell(row=sheet.max_row, column=col).coordinate,
                            )
                        )

        buffer.seek(0)

        # Criar a resposta para o download
        response = HttpResponse(
            buffer,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="RELATORIO-GZH_{data_atual}.xlsx"'
        )

        return response

    except Exception as e:
        return HttpResponse(f"Erro ao gerar insights: {str(e)}", status=500)


# @login_required
# def configurar_cliente_sat():
#     # Configurar o caminho da DLL
#     biblioteca = BibliotecaSAT(settings.SAT_DLL_PATH)

#     # Código de ativação fornecido pelo fabricante
#     codigo_ativacao = '12345678'  # Substitua pelo código real

#     # Criar o cliente SAT
#     cliente = ClienteSATLocal(biblioteca, codigo_ativacao=codigo_ativacao)
#     return cliente

# def emitir_nota_fiscal(request):
#     try:
#         # Configurar o cliente SAT
#         cliente = configurar_cliente_sat()

#         # Consultar o SAT
#         resposta = cliente.consultar_sat()
#         if resposta.codigo == 0:  # Verifica se o código de resposta indica sucesso
#             messages.success(request, 'SAT está em operação: ' + resposta.mensagem)
#         else:
#             messages.error(request, f"Falha ao consultar o SAT: {resposta.mensagem}")

#     except FileNotFoundError as e:
#         messages.error(request, str(e))
#     except Exception as e:
#         messages.error(request, f"Erro inesperado: {str(e)}")

#     return render(request, 'purchase_details.html')
